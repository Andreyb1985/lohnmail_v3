from pathlib import Path
from typing import Iterable

from openpyxl import Workbook


def write_audit_check_xlsx(
    out_path: Path,
    grouped: dict[str, list[Path]],
    email_map: dict[str, str],
    invalid_files: Iterable[Path],
    missing_email_persnr: list[str],
    missing_files_persnr: list[str],
    bundle_details: list[dict] | None = None,
    validation: dict | None = None,
    email_records: dict[str, dict[str, str]] | None = None,
    invalid_pdf_details: list[dict] | None = None,
    pdf_errors_by_persnr: dict[str, list[str]] | None = None,
) -> None:
    def person_name(persnr: str) -> str:
        record = (email_records or {}).get(persnr, {})
        name = str(record.get("Name", "") or "").strip()
        vorname = str(record.get("Vorname", "") or "").strip()
        return ", ".join(part for part in [name, vorname] if part)

    wb = Workbook()

    # ---------- Blatt 1: Übersicht ----------
    ws1 = wb.active
    ws1.title = "Übersicht"

    ws1.append(["PersNr", "Name, Vorname", "Email", "Dateien", "Anzahl", "Status", "Fehler"])

    for persnr in sorted(grouped.keys()):
        files = grouped[persnr]
        email = email_map.get(persnr, "")
        pdf_errors = (pdf_errors_by_persnr or {}).get(persnr, [])
        status = "Fehler" if pdf_errors else ("OK" if email else "Keine E-Mail")
        ws1.append([
            persnr,
            person_name(persnr),
            email,
            ", ".join(p.name for p in files),
            len(files),
            status,
            "; ".join(pdf_errors),
        ])

    for persnr in sorted(missing_files_persnr):
        ws1.append([
            persnr,
            person_name(persnr),
            email_map.get(persnr, ""),
            "",
            0,
            "Keine Dateien",
            "",
        ])

    # ---------- Blatt 2: Ungültige PDF-Dateien ----------
    ws2 = wb.create_sheet("Ungültige Dateien")
    ws2.append(["PersNr", "Dateiname", "Grund"])
    for p in invalid_files:
        ws2.append(["", p.name, "Ungültiger Dateiname"])
    for row in (invalid_pdf_details or []):
        ws2.append([
            row.get("persnr", ""),
            row.get("file", ""),
            row.get("reason", ""),
        ])

    # ---------- Blatt 3: Keine E-Mail ----------
    ws3 = wb.create_sheet("Keine E-Mail")
    ws3.append(["PersNr", "Name, Vorname", "Dateien", "Anzahl"])
    for persnr in sorted(missing_email_persnr):
        files = grouped.get(persnr, [])
        ws3.append([
            persnr,
            person_name(persnr),
            ", ".join(p.name for p in files),
            len(files),
        ])

    # ---------- Blatt 4: Keine Dateien ----------
    ws4 = wb.create_sheet("Keine Dateien")
    ws4.append(["PersNr", "Name, Vorname", "Email"])
    for persnr in sorted(missing_files_persnr):
        ws4.append([persnr, person_name(persnr), email_map.get(persnr, "")])

    # ---------- Blatt 5: Bundle-Prüfung ----------
    ws5 = wb.create_sheet("Bundle-Prüfung")
    ws5.append(["Prüfung", "Wert"])
    if validation:
        order = [
            "total_input_pdf_files",
            "valid_input_pdf_files",
            "invalid_input_pdf_files",
            "unreadable_pdf_files",
            "employees_with_email",
            "employees_without_email",
            "expected_bundle_pdf_files",
            "duplicate_bundle_pdf_files",
            "expected_bundle_pages",
            "actual_bundle_pages",
            "page_check_ok",
        ]
        labels = {
            "total_input_pdf_files": "PDF-Dateien gesamt",
            "valid_input_pdf_files": "Gültige PDF-Dateien",
            "invalid_input_pdf_files": "Ungültige PDF-Dateien",
            "unreadable_pdf_files": "Nicht lesbare PDF-Dateien",
            "employees_with_email": "Mitarbeiter mit E-Mail",
            "employees_without_email": "Mitarbeiter ohne E-Mail",
            "expected_bundle_pdf_files": "PDF-Dateien im Sammel-PDF erwartet",
            "duplicate_bundle_pdf_files": "Exakte PDF-Duplikate im Sammel-PDF ausgeschlossen",
            "expected_bundle_pages": "Seiten im Sammel-PDF erwartet",
            "actual_bundle_pages": "Seiten im Sammel-PDF tatsächlich",
            "page_check_ok": "Seitenprüfung OK",
        }
        for key in order:
            ws5.append([labels[key], validation.get(key)])

    # ---------- Blatt 6: Bundle-Details ----------
    ws6 = wb.create_sheet("Bundle-Details")
    ws6.append([
        "PersNr",
        "Datei",
        "Seiten",
        "Included",
        "Reason",
        "FileHash",
        "DuplicateOf",
    ])
    for row in (bundle_details or []):
        ws6.append([
            row.get("persnr", ""),
            row.get("file", ""),
            row.get("pages", 0),
            row.get("included", ""),
            row.get("reason", ""),
            row.get("file_hash", ""),
            row.get("duplicate_of", ""),
        ])

    # einfache Breiten
    for ws in wb.worksheets:
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                val = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(val))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 80)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
