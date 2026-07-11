# core/excel_io.py
from numbers import Integral, Real
from pathlib import Path
import re

from openpyxl import load_workbook

PERSNR_TEXT_RE = re.compile(r"^(\d{1,5})(?:\.0+)?$")


def normalize_persnr(value) -> str | None:
    if value is None:
        return None

    if isinstance(value, bool):
        return None

    if isinstance(value, Integral):
        digits = str(int(value))
        return digits.zfill(5) if 1 <= len(digits) <= 5 else None

    if isinstance(value, Real):
        if not float(value).is_integer():
            return None
        digits = str(int(value))
        return digits.zfill(5) if 1 <= len(digits) <= 5 else None

    text = str(value).strip()
    if not text:
        return None

    match = PERSNR_TEXT_RE.fullmatch(text)
    if not match:
        return None

    return match.group(1).zfill(5)


def _cell_text(value) -> str:
    text = str(value).strip() if value is not None else ""
    return "" if text.lower() == "nan" else text


def _load_email_records(
    excel_path: Path,
    include_rows_without_email: bool = False,
) -> dict[str, dict[str, str]]:
    if excel_path.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ValueError("Es werden nur Excel-Dateien im Format .xlsx oder .xlsm unterstützt.")

    wb = load_workbook(excel_path, data_only=True)
    ws = wb.active

    headers = {}
    for idx, cell in enumerate(ws[1], start=1):
        val = str(cell.value).strip() if cell.value is not None else ""
        headers[val] = idx

    if "PersNr" not in headers or "Email" not in headers:
        raise ValueError("Excel muss die Spalten 'PersNr' und 'Email' enthalten.")

    persnr_col = headers["PersNr"]
    email_col = headers["Email"]
    name_col = headers.get("Name")
    vorname_col = headers.get("Vorname")

    result: dict[str, dict[str, str]] = {}
    persnr_rows: dict[str, int] = {}
    email_rows: dict[str, tuple[int, str]] = {}

    for row_number, row in enumerate(ws.iter_rows(min_row=2), start=2):
        persnr_raw = row[persnr_col - 1].value
        email_raw = row[email_col - 1].value
        name_raw = row[name_col - 1].value if name_col else None
        vorname_raw = row[vorname_col - 1].value if vorname_col else None

        persnr = normalize_persnr(persnr_raw)
        email = _cell_text(email_raw)
        name = _cell_text(name_raw)
        vorname = _cell_text(vorname_raw)
        email_key = email.lower()

        if not persnr and _cell_text(persnr_raw):
            raise ValueError(
                f"Ungültige PersNr in Excel-Zeile {row_number}: {persnr_raw!r}. "
                "Erlaubt sind nur ganze Zahlen mit maximal 5 Stellen."
            )

        if not persnr:
            continue

        if not email and not include_rows_without_email:
            continue

        if not email and not (name or vorname):
            continue

        if persnr in result:
            first_row = persnr_rows.get(persnr, "?")
            raise ValueError(
                f"Doppelte PersNr {persnr} in Excel-Zeilen {first_row} und {row_number}."
            )

        if email and email_key in email_rows:
            first_row, first_persnr = email_rows[email_key]
            raise ValueError(
                f"Doppelte E-Mail-Adresse {email} in Excel-Zeilen {first_row} und {row_number} "
                f"(PersNr {first_persnr} und {persnr})."
            )

        result[persnr] = {
            "PersNr": persnr,
            "Email": email,
            "Name": name,
            "Vorname": vorname,
        }
        persnr_rows[persnr] = row_number
        if email:
            email_rows[email_key] = (row_number, persnr)

    return result


def load_email_records(excel_path: Path) -> dict[str, dict[str, str]]:
    return _load_email_records(excel_path, include_rows_without_email=True)


def load_email_map(excel_path: Path) -> dict[str, str]:
    records = _load_email_records(excel_path, include_rows_without_email=False)
    return {
        persnr: record.get("Email", "")
        for persnr, record in records.items()
        if record.get("Email", "")
    }
