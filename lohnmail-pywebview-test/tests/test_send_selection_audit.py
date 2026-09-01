from __future__ import annotations

from pathlib import Path
from tempfile import TemporaryDirectory
import unittest
from unittest.mock import patch

import fitz
from openpyxl import Workbook, load_workbook

from core import orchestrator


class SendSelectionAuditTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temp_dir = TemporaryDirectory()
        self.root = Path(self.temp_dir.name)
        self.pdf_dir = self.root / "input"
        self.pdf_dir.mkdir()
        self.output_dir = self.root / "Gesob_Lohn"

    def tearDown(self) -> None:
        self.temp_dir.cleanup()

    def _create_pdf(self, filename: str, label: str) -> None:
        pdf_path = self.pdf_dir / filename
        doc = fitz.open()
        page = doc.new_page(width=420, height=260)
        page.insert_text((36, 90), label, fontsize=24)
        doc.save(pdf_path)
        doc.close()

    def _create_excel(self) -> Path:
        excel_path = self.root / "employees.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["PersNr", "Email", "Name", "Vorname"])
        sheet.append(["00001", "one@example.com", "Eins", "Anna"])
        sheet.append(["00002", "", "Zwei", "Berta"])
        sheet.append(["00003", "three@example.com", "Drei", "Clara"])
        workbook.save(excel_path)
        return excel_path

    @staticmethod
    def _bundle_values(audit_path: Path) -> dict[str, object]:
        workbook = load_workbook(audit_path, data_only=True)
        sheet = workbook["Bundle-Prüfung"]
        return {
            str(label): value
            for label, value in sheet.iter_rows(min_row=2, values_only=True)
            if label
        }

    def test_recipient_selection_does_not_reduce_audit_or_missing_email_bundle(self) -> None:
        self._create_pdf("00001.pdf", "Mitarbeiter 00001")
        self._create_pdf("00002.pdf", "Mitarbeiter 00002 Teil 1")
        self._create_pdf("00002_1.pdf", "Mitarbeiter 00002 Teil 2")
        self._create_pdf("00003.pdf", "Mitarbeiter 00003")
        excel_path = self._create_excel()
        settings = {
            "mail_mode": "smtp",
            "smtp": {},
            "mail_text": {
                "subject": "Abrechnung {monat} {jahr}",
                "body": "Test {persnr}",
                "body_html": "",
            },
            "pdf_password": {"enabled": False},
            "companies": [{"id": "test", "name": "Test GmbH"}],
            "selected_company_id": "test",
            "period": {"mode": "automatic_current_month"},
        }

        with (
            patch.object(orchestrator, "GESOB_DIR", self.output_dir),
            patch.object(orchestrator, "make_run_id", return_value="send_selection_test"),
            patch.object(
                orchestrator,
                "validate_email_records",
                side_effect=lambda records: {
                    persnr: {
                        "code": "missing" if not record.get("Email") else "valid",
                        "sendable": bool(record.get("Email")),
                    }
                    for persnr, record in records.items()
                },
            ),
        ):
            result = orchestrator.action_send(
                pdf_input=self.pdf_dir,
                excel_path=excel_path,
                settings=settings,
                company_id="test",
                dry_run=True,
                allowed_persnr={"00001"},
            )

        self.assertEqual([row["PersNr"] for row in result["table_rows"]], ["00001"])
        self.assertEqual(result["summary"]["total_pdf_files"], 4)
        self.assertEqual(result["summary"]["unique_persnr_count"], 3)
        self.assertEqual(result["summary"]["missing_email_count"], 1)
        self.assertEqual(result["summary"]["missing_pdf_count"], 2)
        self.assertEqual(result["summary"]["expected_bundle_pages"], 2)
        self.assertEqual(result["summary"]["actual_bundle_pages"], 2)
        self.assertTrue(result["summary"]["page_check_ok"])

        missing_pdf_path = result["missing_pdf_path"]
        self.assertIsNotNone(missing_pdf_path)
        missing_pdf = fitz.open(missing_pdf_path)
        try:
            self.assertEqual(missing_pdf.page_count, 2)
        finally:
            missing_pdf.close()

        bundle = self._bundle_values(result["audit_path"])
        self.assertEqual(bundle["PDF-Dateien gesamt"], 4)
        self.assertEqual(bundle["Mitarbeiter mit E-Mail"], 2)
        self.assertEqual(bundle["Mitarbeiter ohne E-Mail"], 1)
        self.assertEqual(bundle["PDF-Dateien im Sammel-PDF erwartet"], 2)
        self.assertEqual(bundle["Seiten im Sammel-PDF erwartet"], 2)
        self.assertEqual(bundle["Seiten im Sammel-PDF tatsächlich"], 2)
        self.assertTrue(bundle["Gesamtprüfung OK"])


if __name__ == "__main__":
    unittest.main()
