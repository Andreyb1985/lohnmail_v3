# core/excel_io.py
from datetime import date, datetime
from numbers import Integral, Real
from pathlib import Path
import re

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

PERSNR_TEXT_RE = re.compile(r"^(\d{1,5})(?:\.0+)?$")


def _parse_short_birth_date(value: str) -> date | None:
    """Parse DATEV-style DDMMYY dates using a plausible employee age."""
    if not re.fullmatch(r"\d{6}", value):
        return None

    today = date.today()
    day = int(value[:2])
    month = int(value[2:4])
    short_year = int(value[4:])
    candidates: list[date] = []

    for year in (2000 + short_year, 1900 + short_year):
        try:
            candidate = date(year, month, day)
        except ValueError:
            continue
        age = today.year - candidate.year - (
            (today.month, today.day) < (candidate.month, candidate.day)
        )
        if 14 <= age <= 110:
            candidates.append(candidate)

    return max(candidates) if candidates else None


def normalize_persnr(value) -> str | None:
    if value is None:
        return None

    if isinstance(value, bool):
        return None

    if isinstance(value, Integral):
        digits = str(int(value))
        return digits.zfill(5) if 1 <= len(digits) <= 5 else None

    if isinstance(value, Real):
        if not float(value).is_integer():
            return None
        digits = str(int(value))
        return digits.zfill(5) if 1 <= len(digits) <= 5 else None

    text = str(value).strip()
    if not text:
        return None

    match = PERSNR_TEXT_RE.fullmatch(text)
    if not match:
        return None

    return match.group(1).zfill(5)


def _cell_text(value) -> str:
    text = str(value).strip() if value is not None else ""
    return "" if text.lower() == "nan" else text


def normalize_birth_date(value, excel_epoch=None) -> str:
    """Return a birth date in the password-safe DDMMYYYY format."""
    parsed_date: date | None = None

    if value is None or isinstance(value, bool):
        return ""
    if isinstance(value, datetime):
        parsed_date = value.date()
    elif isinstance(value, date):
        parsed_date = value
    elif isinstance(value, (Integral, Real)):
        digits = str(int(value)) if float(value).is_integer() else ""
        if len(digits) == 6:
            parsed_date = _parse_short_birth_date(digits)
        elif len(digits) in {7, 8}:
            try:
                parsed_date = datetime.strptime(digits.zfill(8), "%d%m%Y").date()
            except ValueError:
                parsed_date = None
        if parsed_date is None:
            try:
                converted = from_excel(value, epoch=excel_epoch) if excel_epoch else from_excel(value)
                parsed_date = converted.date() if isinstance(converted, datetime) else converted
            except (TypeError, ValueError, OverflowError):
                parsed_date = None
    else:
        text = _cell_text(value)
        if len(text) == 6:
            parsed_date = _parse_short_birth_date(text)
        for date_format in ("%d.%m.%Y", "%d/%m/%Y", "%Y-%m-%d", "%d%m%Y"):
            if parsed_date is not None:
                break
            try:
                parsed_date = datetime.strptime(text, date_format).date()
                break
            except ValueError:
                continue

    return parsed_date.strftime("%d%m%Y") if isinstance(parsed_date, date) else ""


def _load_email_records(
    excel_path: Path,
    include_rows_without_email: bool = False,
) -> dict[str, dict[str, str]]:
    if excel_path.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ValueError("Es werden nur Excel-Dateien im Format .xlsx oder .xlsm unterstützt.")

    wb = load_workbook(excel_path, data_only=True)
    ws = wb.active

    headers = {}
    for idx, cell in enumerate(ws[1], start=1):
        val = str(cell.value).strip() if cell.value is not None else ""
        headers[val] = idx

    if "PersNr" not in headers or "Email" not in headers:
        raise ValueError("Excel muss die Spalten 'PersNr' und 'Email' enthalten.")

    persnr_col = headers["PersNr"]
    email_col = headers["Email"]
    name_col = headers.get("Name")
    vorname_col = headers.get("Vorname")
    birth_date_col = headers.get("Geburtsdatum") or headers.get("Gebursdatum")

    result: dict[str, dict[str, str]] = {}
    persnr_rows: dict[str, int] = {}

    for row_number, row in enumerate(ws.iter_rows(min_row=2), start=2):
        persnr_raw = row[persnr_col - 1].value
        email_raw = row[email_col - 1].value
        name_raw = row[name_col - 1].value if name_col else None
        vorname_raw = row[vorname_col - 1].value if vorname_col else None
        birth_date_raw = row[birth_date_col - 1].value if birth_date_col else None

        persnr = normalize_persnr(persnr_raw)
        email = _cell_text(email_raw)
        name = _cell_text(name_raw)
        vorname = _cell_text(vorname_raw)
        birth_date = normalize_birth_date(birth_date_raw, wb.epoch)

        if not persnr and _cell_text(persnr_raw):
            raise ValueError(
                f"Ungültige PersNr in Excel-Zeile {row_number}: {persnr_raw!r}. "
                "Erlaubt sind nur ganze Zahlen mit maximal 5 Stellen."
            )

        if not persnr:
            continue

        if not email and not include_rows_without_email:
            continue

        if not email and not (name or vorname):
            continue

        if persnr in result:
            first_row = persnr_rows.get(persnr, "?")
            raise ValueError(
                f"Doppelte PersNr {persnr} in Excel-Zeilen {first_row} und {row_number}."
            )

        result[persnr] = {
            "PersNr": persnr,
            "Email": email,
            "Name": name,
            "Vorname": vorname,
            "Geburtsdatum": birth_date,
            "ExcelRow": row_number,
        }
        persnr_rows[persnr] = row_number

    return result


def load_email_records(excel_path: Path) -> dict[str, dict[str, str]]:
    return _load_email_records(excel_path, include_rows_without_email=True)


def load_email_map(excel_path: Path) -> dict[str, str]:
    records = _load_email_records(excel_path, include_rows_without_email=False)
    return {
        persnr: record.get("Email", "")
        for persnr, record in records.items()
        if record.get("Email", "")
    }
